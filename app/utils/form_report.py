"""
form_report.py — report utilities for form results.

Functions:
    send_form_report()        — verbose per-question breakdown (existing /form results)
    send_form_summary()       — compact summary + Excel link (/form report)
    form_ai_analysis()        — LLM analysis of submitted answers (/form ai)
"""
import io
import logging
import os
import tempfile
from collections import Counter
from datetime import datetime

import openai
from sqlalchemy.orm import Session

from app.config import get_settings
from app.db import models
from app.whatsapp.client import WahaClient

logger = logging.getLogger(__name__)
wa = WahaClient()


def _normalize_jid(jid: str | None) -> str:
    if not jid:
        return ""
    return jid.split("@", 1)[0].lstrip("+")


def _get_submission_student_name(submission: models.FormSubmission, db: Session) -> str:
    if submission.student_id:
        student = submission.student or db.query(models.Student).get(submission.student_id)
        if student:
            return student.name

    normalized_jid = _normalize_jid(submission.respondent_jid)
    for contact in db.query(models.KnownContact).all():
        if _normalize_jid(contact.jid) == normalized_jid:
            # child_name now on KCG; return first non-null
            kcg = db.query(models.KnownContactGroup).filter(
                models.KnownContactGroup.contact_jid == contact.jid,
                models.KnownContactGroup.child_name.isnot(None),
                models.KnownContactGroup.active == True,
            ).first()
            return kcg.child_name if kcg else ""
    return ""


def _get_classroom_name(classroom_id: int | None, db: Session) -> str:
    if not classroom_id:
        return ""
    classroom = db.query(models.Classroom).get(classroom_id)
    if not classroom:
        return ""
    return classroom.display_name or classroom.name


def _get_submission_classroom(
    submission: models.FormSubmission,
    audience_classroom_ids: list[int],
    db: Session,
) -> str:
    """Resolve the classroom(s) a submission belongs to.

    Matters for forms whose audience spans several classrooms — otherwise every
    row looks alike. Falls back through: linked student → group membership →
    single-classroom audience.
    """
    try:
        if submission.student_id:
            student = submission.student or db.query(models.Student).get(submission.student_id)
            if student and student.classroom_id:
                return _get_classroom_name(student.classroom_id, db)

        normalized_jid = _normalize_jid(submission.respondent_jid)
        contact = None
        for c in db.query(models.KnownContact).all():
            if _normalize_jid(c.jid) == normalized_jid:
                contact = c
                break
        if not contact:
            # {phone}@c.us respondent vs {lid}@lid contact — bridge on phone.
            from app.utils.jid_utils import safe_phone

            phone = safe_phone(submission.respondent_jid or "", None)
            if phone:
                contact = db.query(models.KnownContact).filter_by(phone=phone).first()
        if contact:
            q = db.query(models.KnownContactGroup).filter_by(contact_jid=contact.jid, active=True)
            if audience_classroom_ids:
                q = q.filter(models.KnownContactGroup.classroom_id.in_(audience_classroom_ids))
            names = [
                name
                for name in (
                    _get_classroom_name(g.classroom_id, db)
                    for g in q.order_by(models.KnownContactGroup.classroom_id).all()
                )
                if name
            ]
            if names:
                return ", ".join(names)

        if len(audience_classroom_ids) == 1:
            return _get_classroom_name(audience_classroom_ids[0], db)

        return ""
    except Exception:
        return ""


def _parents_in_audience(classroom_ids: list[int], db: Session) -> list[models.Parent]:
    """Active registered parents with at least one student in the given classrooms."""
    if not classroom_ids:
        return []
    students = db.query(models.Student).filter(
        models.Student.classroom_id.in_(classroom_ids)
    ).all()
    parent_ids = {s.parent_id for s in students if s.parent_id}
    if not parent_ids:
        return []
    return db.query(models.Parent).filter(
        models.Parent.id.in_(parent_ids),
        models.Parent.is_active == True,
    ).all()


_PURPOSE_LABELS = {
    "intake": "Datos / Perfil",
    "survey": "Encuesta",
    "event_registration": "Registro de evento",
    "volunteer_signup": "Voluntariado",
}

_STATUS_ICONS = {"draft": "📝", "open": "✅", "closed": "🔒", "archived": "🗄️"}


# ── Excel export ───────────────────────────────────────────────────────────────

def _build_form_excel(form: models.Form, db: Session) -> bytes:
    """Build Excel workbook bytes from all submissions."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    questions = (
        db.query(models.FormQuestion)
        .filter_by(form_id=form.id)
        .order_by(models.FormQuestion.order)
        .all()
    )
    submissions = db.query(models.FormSubmission).filter_by(form_id=form.id).all()
    audience_classroom_ids = [
        a.classroom_id for a in db.query(models.FormAudience).filter_by(form_id=form.id).all()
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Respuestas"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["Padre/Madre", "Estudiante", "Salón", "Estado", "Enviado", "Iniciado"]
    for q in questions:
        headers.append(q.text[:60].replace("\n", " "))
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for sub in submissions:
        answers_map = {
            a.question_id: a.value
            for a in db.query(models.FormAnswer).filter_by(submission_id=sub.id).all()
        }
        student_name = _get_submission_student_name(sub, db)
        row = [
            sub.respondent_name,
            student_name,
            _get_submission_classroom(sub, audience_classroom_ids, db),
            sub.status,
            sub.submitted_at.strftime("%Y-%m-%d %H:%M") if sub.submitted_at else "",
            sub.started_at.strftime("%Y-%m-%d %H:%M") if sub.started_at else "",
        ]
        for q in questions:
            val = answers_map.get(q.id)
            if val == "yes":
                val = "Sí"
            elif val == "no":
                val = "No"
            row.append(val or "")
        ws.append(row)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_excel_url(form: models.Form, db: Session) -> str:
    """Build Excel, upload to S3, return shortened presigned URL."""
    from app.utils.s3_upload import upload_file_to_s3, generate_presigned_url
    from app.utils.helpers import shorten_url

    content = _build_form_excel(form, db)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(content)
        tmp_path = f.name

    try:
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        s3_key = f"forms/form_{form.id}_{timestamp}.xlsx"
        upload_file_to_s3(tmp_path, s3_key)
        presigned = generate_presigned_url(s3_key)
        return shorten_url(presigned)
    finally:
        os.unlink(tmp_path)


# ── Compact summary (/form report) ────────────────────────────────────────────

def send_form_summary(form: models.Form, chat_id: str, db: Session):
    """Send a compact summary with counts + CSV download link."""
    audience_rows = db.query(models.FormAudience).filter_by(form_id=form.id).all()
    classroom_ids = [a.classroom_id for a in audience_rows]
    audience_parents = _parents_in_audience(classroom_ids, db)
    total_parents = len(audience_parents)

    submitted_subs = db.query(models.FormSubmission).filter_by(
        form_id=form.id, status="submitted"
    ).all()
    submitted = len(submitted_subs)
    in_progress = db.query(models.FormSubmission).filter_by(
        form_id=form.id, status="in_progress"
    ).count()
    pending = max(0, total_parents - submitted - in_progress)

    submitted_jids = {s.respondent_jid for s in submitted_subs}
    non_respondents = [p for p in audience_parents if p.whatsapp_jid not in submitted_jids]

    purpose_label = _PURPOSE_LABELS.get(form.purpose, form.purpose)
    status_icon = _STATUS_ICONS.get(form.status, "❓")

    msg_lines = [
        f"📊 *{form.title}*",
        f"{status_icon} {form.status} | {purpose_label}",
        "",
        f"👥 Audiencia: *{total_parents}* padres",
        f"✅ Respondieron: *{submitted}*",
        f"⏳ En progreso: *{in_progress}*",
        f"❌ Pendientes: *{pending}*",
    ]

    if non_respondents:
        names = [f"{p.first_name} {p.last_name}" for p in non_respondents[:15]]
        if len(non_respondents) > 15:
            names.append(f"_y {len(non_respondents) - 15} más_")
        msg_lines.append("\n_Sin responder:_\n  — " + "\n  — ".join(names))

    if submitted == 0:
        msg_lines.append("\n_No hay respuestas enviadas aún._")
        wa.send_text(chat_id, "\n".join(msg_lines))
        return

    # Generate Excel and append link
    try:
        excel_url = _generate_excel_url(form, db)
        msg_lines.append(f"\n📊 *Excel completo:*\n{excel_url}")
    except Exception as e:
        logger.error("Excel generation failed form_id=%d: %s", form.id, e)
        msg_lines.append("\n_(Error al generar Excel)_")

    wa.send_text(chat_id, "\n".join(msg_lines))
    logger.info("FORM summary sent form_id=%d submitted=%d", form.id, submitted)


# ── AI analysis (/form ai) ─────────────────────────────────────────────────────

def _build_form_context(form: models.Form, db: Session) -> str:
    """Build a structured text context of all submitted answers for the LLM."""
    questions = (
        db.query(models.FormQuestion)
        .filter_by(form_id=form.id)
        .order_by(models.FormQuestion.order)
        .all()
    )
    submissions = db.query(models.FormSubmission).filter_by(
        form_id=form.id, status="submitted"
    ).all()

    # Pre-load all answers
    sub_ids = [s.id for s in submissions]
    all_answers = db.query(models.FormAnswer).filter(
        models.FormAnswer.submission_id.in_(sub_ids)
    ).all() if sub_ids else []
    answers_by_sub: dict = {}
    for a in all_answers:
        answers_by_sub.setdefault(a.submission_id, {})[a.question_id] = a.value

    lines = [
        f'Formulario: "{form.title}"',
        f"Propósito: {_PURPOSE_LABELS.get(form.purpose, form.purpose)}",
        f"Total de respuestas enviadas: {len(submissions)}",
        "",
        "Preguntas:",
    ]
    for i, q in enumerate(questions, 1):
        lines.append(f"  {i}. {q.text}")

    # Pre-computed aggregates per question (accurate counts for the LLM)
    lines.append("\nESTADÍSTICAS AGREGADAS POR PREGUNTA (valores exactos calculados en código):")
    for q in questions:
        values = []
        for sub in submissions:
            val = answers_by_sub.get(sub.id, {}).get(q.id)
            if val:
                display = "Sí" if val == "yes" else ("No" if val == "no" else val)
                values.append(display)
        if not values:
            continue
        counter = Counter(values)
        lines.append(f"\n  Pregunta: {q.text}")
        lines.append(f"  Total respondieron: {len(values)} de {len(submissions)}")
        for val, count in counter.most_common():
            lines.append(f"    • {val}: {count}")

    lines.append("\nRESPUESTAS INDIVIDUALES:")
    for sub in submissions:
        answers_map = answers_by_sub.get(sub.id, {})
        lines.append(f"\n• {sub.respondent_name}:")
        for q in questions:
            val = answers_map.get(q.id)
            if val is None:
                continue
            display = "Sí" if val == "yes" else ("No" if val == "no" else val)
            lines.append(f"    - {q.text}: {display}")

    return "\n".join(lines)


async def form_ai_analysis(form: models.Form, question: str, chat_id: str, db: Session):
    """Send submitted answers to the LLM and answer a specific analysis question."""
    settings = get_settings()

    submissions_count = db.query(models.FormSubmission).filter_by(
        form_id=form.id, status="submitted"
    ).count()

    if submissions_count == 0:
        wa.send_text(chat_id, "❌ No hay respuestas enviadas para analizar.")
        return

    wa.send_text(chat_id, f"🤖 Analizando {submissions_count} respuestas...")

    context = _build_form_context(form, db)

    client = openai.OpenAI(api_key=settings.openai_api_key)
    response = client.chat.completions.create(
        model=settings.openai_model,
        messages=[
            {
                "role": "system",
                "content": (
                    "Eres un asistente de análisis de datos escolares. "
                    "Se te proporcionan las respuestas de un formulario escolar. "
                    "Responde la pregunta del usuario basándote ESTRICTAMENTE en los datos proporcionados. "
                    "Responde en español, de forma clara y concisa. "
                    "Si los datos no son suficientes para responder, dilo claramente. "
                    "El contexto incluye una sección 'ESTADÍSTICAS AGREGADAS' con conteos exactos calculados en código — "
                    "SIEMPRE usa esos números para responder preguntas de conteo o agrupación, nunca los recalcules tú mismo."
                ),
            },
            {
                "role": "user",
                "content": f"Datos del formulario:\n\n{context}\n\nPregunta: {question}",
            },
        ],
        temperature=0.2,
        max_completion_tokens=3000,
    )

    reply = response.choices[0].message.content
    wa.send_text(chat_id, reply)
    logger.info("FORM ai_analysis form_id=%d question_len=%d", form.id, len(question))


# ── Verbose per-question report (existing /form results) ──────────────────────

def send_form_report(form: models.Form, chat_id: str, db: Session):
    """Generate and send a verbose text report of form results to chat_id."""
    questions = (
        db.query(models.FormQuestion)
        .filter_by(form_id=form.id)
        .order_by(models.FormQuestion.order)
        .all()
    )

    submissions = (
        db.query(models.FormSubmission)
        .filter_by(form_id=form.id, status="submitted")
        .all()
    )

    audience_rows = db.query(models.FormAudience).filter_by(form_id=form.id).all()
    classroom_ids = [a.classroom_id for a in audience_rows]
    audience_parents = _parents_in_audience(classroom_ids, db)
    total_parents = len(audience_parents)

    sub_count = len(submissions)
    pending = total_parents - sub_count if total_parents >= sub_count else 0

    purpose_label = _PURPOSE_LABELS.get(form.purpose, form.purpose)
    status_icon = _STATUS_ICONS.get(form.status, "❓")

    header = (
        f"📊 *Resultados: {form.title}*\n"
        f"{status_icon} {form.status} | {purpose_label}\n\n"
        f"Audiencia: *{total_parents}* padres\n"
        f"Respondieron: *{sub_count}* | Pendientes: *{pending}*"
    )
    wa.send_text(chat_id, header)

    if not submissions:
        wa.send_text(chat_id, "_No hay respuestas enviadas aún._")
        return

    submission_ids = [s.id for s in submissions]

    for q in questions:
        answers = (
            db.query(models.FormAnswer)
            .filter(
                models.FormAnswer.submission_id.in_(submission_ids),
                models.FormAnswer.question_id == q.id,
            )
            .all()
        )

        answered_count = sum(1 for a in answers if a.value is not None)
        skipped = sub_count - answered_count

        lines = [f"❓ *{q.text}*\n   ({answered_count} respuestas, {skipped} sin responder)"]

        if q.type == "yes_no":
            values = [a.value for a in answers if a.value]
            lines.append(f"   ✅ Sí: *{values.count('yes')}*  ❌ No: *{values.count('no')}*")

        elif q.type == "single_choice":
            values = [a.value for a in answers if a.value]
            counter = Counter(values)
            for opt in (q.options or []):
                count = counter.get(opt, 0)
                bar = "█" * min(count, 20)
                lines.append(f"   • {opt}: *{count}* {bar}")

        elif q.type == "text":
            text_answers = [a.value for a in answers if a.value]
            shown = text_answers[:15]
            for ans in shown:
                lines.append(f"   — {ans}")
            if len(text_answers) > 15:
                lines.append(f"   _... y {len(text_answers) - 15} más_")

        wa.send_text(chat_id, "\n".join(lines))

    submitted_jids = {s.respondent_jid for s in submissions}
    non_respondents = [p for p in audience_parents if p.whatsapp_jid not in submitted_jids]

    if non_respondents:
        names = [f"{p.first_name} {p.last_name}" for p in non_respondents]
        pending_lines = ["⏳ *Padres que no han respondido:*"]
        for name in names:
            pending_lines.append(f"  — {name}")
        wa.send_text(chat_id, "\n".join(pending_lines))

    logger.info("FORM report sent form_id=%d submissions=%d", form.id, sub_count)
