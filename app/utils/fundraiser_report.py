"""
Fundraiser Excel report generator.

Generates an .xlsx file, uploads to S3, and returns a presigned download URL.

For variable (catalog) fundraisers: one column per product with quantities, totals row at end.
For fixed fundraisers: simple rows with amounts, totals row at end.
"""
import io
import logging
import os
import tempfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.db import models

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
_TOTAL_FONT = Font(bold=True)


def generate_fundraiser_excel_url(
    fundraiser: models.Fundraiser,
    payments: list[models.Payment],
    db: Session,
) -> str:
    """Build Excel workbook, upload to S3, return presigned URL."""
    from app.utils.s3_upload import upload_file_to_s3, generate_presigned_url
    from app.utils.helpers import shorten_url

    content = _build_excel(fundraiser, payments, db)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(content)
        tmp_path = f.name

    try:
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        s3_key = f"fundraisers/fundraiser_{fundraiser.id}_{timestamp}.xlsx"
        upload_file_to_s3(tmp_path, s3_key)
        presigned = generate_presigned_url(s3_key)
        return shorten_url(presigned)
    finally:
        os.unlink(tmp_path)


def _build_excel(
    fundraiser: models.Fundraiser,
    payments: list[models.Payment],
    db: Session,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    if fundraiser.type == "variable":
        _write_variable(ws, fundraiser, payments, db)
    else:
        _write_fixed(ws, payments, db)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _apply_header(ws, headers: list[str]):
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def _normalize_jid(jid: str | None) -> str:
    if not jid:
        return ""
    return jid.split("@", 1)[0].lstrip("+")


def _get_classroom_name(classroom_id: int | None, db: Session) -> str:
    if not classroom_id:
        return ""
    classroom = db.query(models.Classroom).get(classroom_id)
    return classroom.name if classroom else ""


def _find_parent_by_jid(jid: str, db: Session) -> models.Parent | None:
    normalized_jid = _normalize_jid(jid)
    for parent in db.query(models.Parent).filter(models.Parent.is_active == True).all():
        if _normalize_jid(parent.whatsapp_jid) == normalized_jid:
            return parent
    return None


def _find_known_contact_by_jid(jid: str, db: Session) -> models.KnownContact | None:
    normalized_jid = _normalize_jid(jid)
    for contact in db.query(models.KnownContact).all():
        if _normalize_jid(contact.jid) == normalized_jid:
            return contact
    return None


def _get_payment_group_name(payment: models.Payment, fundraiser: models.Fundraiser, db: Session) -> str:
    """Resolve the most relevant classroom/group name for a payment."""
    try:
        parent = _find_parent_by_jid(payment.payer_jid, db)
        if parent:
            student = (
                db.query(models.Student)
                .filter_by(parent_id=parent.id)
                .filter(models.Student.name.ilike(payment.child_name or ""))
                .first()
            )
            if student and student.classroom_id:
                return _get_classroom_name(student.classroom_id, db)

        contact = _find_known_contact_by_jid(payment.payer_jid, db)
        if contact:
            query = db.query(models.KnownContactGroup).filter_by(
                contact_jid=contact.jid,
                active=True,
            )
            if fundraiser.audience_classroom_ids:
                query = query.filter(
                    models.KnownContactGroup.classroom_id.in_(fundraiser.audience_classroom_ids)
                )
            group_names = [
                name
                for name in (
                    _get_classroom_name(group.classroom_id, db)
                    for group in query.order_by(models.KnownContactGroup.classroom_id).all()
                )
                if name
            ]
            if group_names:
                return ", ".join(group_names)

        if fundraiser.audience_classroom_ids and len(fundraiser.audience_classroom_ids) == 1:
            return _get_classroom_name(fundraiser.audience_classroom_ids[0], db)

        return ""
    except Exception:
        return ""


def _get_student_classroom(payment, db) -> str:
    """Backward-compatible wrapper used by the Excel row builders."""
    fundraiser = payment.fundraiser or db.query(models.Fundraiser).get(payment.fundraiser_id)
    if not fundraiser:
        return ""
    return _get_payment_group_name(payment, fundraiser, db)


def _write_variable(ws, fundraiser, payments, db):
    """One column per product; quantities per payer; totals row at bottom."""
    products = (
        db.query(models.FundraiserProduct)
        .filter_by(fundraiser_id=fundraiser.id)
        .order_by(models.FundraiserProduct.sort_order)
        .all()
    )

    all_items = (
        db.query(models.OrderItem)
        .filter(models.OrderItem.payment_id.in_([p.id for p in payments]))
        .all()
    )
    items_by_payment: dict[int, dict[int, int]] = {}
    for item in all_items:
        items_by_payment.setdefault(item.payment_id, {})[item.product_id] = item.quantity

    headers = ["Nombre", "Estudiante", "Salón"] + [p.name for p in products] + ["Total ($)", "Fecha"]
    _apply_header(ws, headers)

    product_totals = {p.id: 0 for p in products}
    grand_total = 0.0

    for payment in payments:
        order = items_by_payment.get(payment.id, {})
        quantities = []
        for p in products:
            qty = order.get(p.id, 0)
            product_totals[p.id] += qty
            quantities.append(qty if qty else "")

        try:
            amount = float(payment.amount) if payment.amount else 0.0
        except (ValueError, TypeError):
            amount = 0.0
        grand_total += amount

        classroom_name = _get_student_classroom(payment, db)
        date_str = payment.submitted_at.strftime("%Y-%m-%d %H:%M") if payment.submitted_at else ""
        ws.append(
            [payment.payer_name, payment.child_name or "", classroom_name]
            + quantities
            + [round(amount, 2) if amount else "", date_str]
        )

    # Totals row
    totals_row = (
        ["TOTAL", "", ""]
        + [product_totals[p.id] for p in products]
        + [round(grand_total, 2), ""]
    )
    ws.append(totals_row)
    total_row_idx = ws.max_row
    for col_idx in range(1, len(totals_row) + 1):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL

    _autofit(ws)


def _write_fixed(ws, payments, db):
    """Simple name/student/amount table for fixed fundraisers."""
    _apply_header(ws, ["Nombre", "Estudiante", "Salón", "Monto ($)", "Estado", "Fecha"])

    grand_total = 0.0

    for payment in payments:
        try:
            amount = float(payment.amount) if payment.amount else 0.0
        except (ValueError, TypeError):
            amount = 0.0
        grand_total += amount

        classroom_name = _get_student_classroom(payment, db)
        status = "Confirmado" if payment.status == "confirmed" else (
            "Por revisar" if payment.status == "flagged" else payment.status
        )
        date_str = payment.submitted_at.strftime("%Y-%m-%d %H:%M") if payment.submitted_at else ""
        ws.append([
            payment.payer_name,
            payment.child_name or "",
            classroom_name,
            round(amount, 2) if amount else "",
            status,
            date_str,
        ])

    # Totals row
    ws.append(["TOTAL", "", "", round(grand_total, 2), "", ""])
    total_row_idx = ws.max_row
    for col_idx in range(1, 7):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL

    _autofit(ws)
